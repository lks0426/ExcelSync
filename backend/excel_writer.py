#!/usr/bin/env python3
"""
ExcelSync项目的Excel写入模块
将API数据写入mapping.xlsx的D列特定单元格
"""

import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, Optional
import logging
import json
from pathlib import Path

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    处理将API数据写入Excel试算表的特定单元格
    只处理E列有值的D列34个特定单元格
    """
    
    def __init__(self, excel_path: str = "mapping.xlsx", sheet_name: str = "A社貼り付けBS"):
        """
        初始化Excel写入器
        
        参数:
            excel_path: Excel文件路径
            sheet_name: 要写入的工作表名称
        """
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        
    def load_workbook(self):
        """加载Excel工作簿并选择工作表"""
        try:
            self.workbook = load_workbook(self.excel_path)
            self.worksheet = self.workbook[self.sheet_name]
            logger.info(f"成功加载工作簿: {self.excel_path}")
            logger.info(f"已选择工作表: {self.sheet_name}")
        except FileNotFoundError:
            logger.error(f"找不到Excel文件: {self.excel_path}")
            raise
        except KeyError:
            logger.error(f"工作簿中找不到工作表 '{self.sheet_name}'")
            raise
        except Exception as e:
            logger.error(f"加载工作簿时出错: {str(e)}")
            raise
    
    def write_data(self, api_data: Dict[str, Any], mapping: Dict[str, str]) -> Dict[str, str]:
        """
        根据映射将API数据写入特定单元格
        
        参数:
            api_data: 包含API响应数据的字典
            mapping: API字段名到单元格位置的映射字典
            
        返回:
            每个字段的写入状态字典
        """
        if not self.workbook:
            self.load_workbook()
            
        write_status = {}
        
        for api_field, cell_location in mapping.items():
            try:
                value = api_data.get(api_field)
                
                if value is None:
                    logger.warning(f"API数据中找不到字段 '{api_field}'")
                    write_status[api_field] = "missing"
                    continue
                
                # 写入单元格
                self.worksheet[cell_location] = value
                logger.info(f"已将 {api_field} = {value} 写入单元格 {cell_location}")
                write_status[api_field] = "success"
                
            except Exception as e:
                logger.error(f"写入 {api_field} 到 {cell_location} 时出错: {str(e)}")
                write_status[api_field] = f"error: {str(e)}"
        
        return write_status
    
    def save_workbook(self, output_path: Optional[str] = None):
        """
        保存工作簿
        
        参数:
            output_path: 可选的保存路径。如果为None，则覆盖原文件。
        """
        save_path = output_path or self.excel_path
        
        try:
            self.workbook.save(save_path)
            logger.info(f"工作簿已成功保存到: {save_path}")
        except Exception as e:
            logger.error(f"保存工作簿时出错: {str(e)}")
            raise
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            logger.info("工作簿已关闭")
    
    def validate_mapping(self, mapping: Dict[str, str]) -> bool:
        """
        验证映射中的所有单元格都存在且E列有值
        
        参数:
            mapping: API字段名到单元格位置的映射字典
            
        返回:
            布尔值表示映射是否有效
        """
        if not self.workbook:
            self.load_workbook()
        
        valid = True
        
        for api_field, cell_location in mapping.items():
            try:
                # 检查单元格是否存在
                cell = self.worksheet[cell_location]
                
                # 获取对应的E列单元格
                row = int(cell_location[1:])
                e_cell = f"E{row}"
                e_value = self.worksheet[e_cell].value
                
                if e_value is None or (isinstance(e_value, str) and e_value.strip() == ""):
                    logger.warning(f"单元格 {e_cell} (对应 {api_field}) 没有值")
                    valid = False
                else:
                    logger.debug(f"单元格 {e_cell} 有值: {e_value}")
                    
            except Exception as e:
                logger.error(f"验证单元格 {cell_location} 时出错: {str(e)}")
                valid = False
        
        return valid
    
    def process_api_data(self, api_data: Dict[str, Any], mapping: Dict[str, str], output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        完整的API数据处理和写入Excel的工作流程
        
        参数:
            api_data: 包含API响应数据的字典
            mapping: API字段名到单元格位置的映射字典
            output_path: 可选的输出文件路径。如果为None，则覆盖原文件。
            
        返回:
            包含状态和详细信息的处理结果
        """
        result = {
            "status": "started",
            "mapping_valid": False,
            "write_status": {},
            "errors": []
        }
        
        try:
            # 加载工作簿
            self.load_workbook()
            
            # 验证映射
            result["mapping_valid"] = self.validate_mapping(mapping)
            
            if not result["mapping_valid"]:
                result["errors"].append("映射验证失败")
                result["status"] = "validation_failed"
                return result
            
            # 写入数据
            result["write_status"] = self.write_data(api_data, mapping)
            
            # 检查是否所有写入都成功
            failed_writes = [field for field, status in result["write_status"].items() 
                           if status != "success"]
            
            if failed_writes:
                result["errors"].append(f"写入失败的字段: {failed_writes}")
                result["status"] = "partial_success"
            else:
                result["status"] = "success"
            
            # 保存工作簿
            self.save_workbook(output_path)
            
        except Exception as e:
            result["errors"].append(str(e))
            result["status"] = "error"
            logger.error(f"处理API数据时出错: {str(e)}")
        
        finally:
            self.close()
        
        return result


# 使用示例
if __name__ == "__main__":
    from mapping_config import TRIAL_BALANCE_MAPPING
    
    # 示例API数据
    sample_api_data = {
        "cash": 1000000,
        "ordinary_deposits": 5000000,
        "cash_and_deposits_total": 6000000,
        "accounts_receivable": 2000000,
        # ... 其他字段
    }
    
    writer = ExcelWriter()
    result = writer.process_api_data(sample_api_data, TRIAL_BALANCE_MAPPING)
    print(json.dumps(result, indent=2, ensure_ascii=False))